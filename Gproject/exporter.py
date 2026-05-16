"""
Sprint 4 — Export Service.
Generates standard-output Excel and PDF deliverables for an approved case
(US4 from the proposal).
"""
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _fmt(v):
    return "—" if v in (None, "") else str(v)


def build_excel(case, documents, fields_by_doc, audit_rows) -> bytes:
    """Build an Excel workbook for a case and return the bytes."""
    wb = Workbook()

    # Sheet 1 — Case summary
    ws = wb.active
    ws.title = "Case Summary"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55

    title = ws.cell(row=1, column=1, value=f"LogiFlow — Case #{case.case_id} Export")
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.fill = HEADER_FILL
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 24

    rows = [
        ("Customer", case.customer),
        ("Service Type", case.service_type),
        ("Status", case.status),
        ("Created At", case.created_at.isoformat() if case.created_at else "—"),
        ("Created By (user_id)", case.created_by),
        ("Documents Count", len(documents)),
        ("Generated At", datetime.utcnow().isoformat(timespec="seconds") + "Z"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        a = ws.cell(row=i, column=1, value=k)
        b = ws.cell(row=i, column=2, value=_fmt(v))
        a.font = Font(bold=True)
        a.border = BORDER
        b.border = BORDER

    # Sheet 2 — Approved fields per document
    ws2 = wb.create_sheet("Approved Fields")
    headers = ["Doc ID", "File", "Field", "Extracted Value", "Approved Value", "Confidence", "Approved At"]
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    for col, w in enumerate([8, 30, 20, 28, 28, 12, 22], start=1):
        ws2.column_dimensions[chr(64 + col)].width = w

    r = 2
    for d in documents:
        fields = fields_by_doc.get(d.doc_id, [])
        for f in fields:
            ws2.cell(row=r, column=1, value=d.doc_id)
            ws2.cell(row=r, column=2, value=d.file_path.split("/")[-1])
            ws2.cell(row=r, column=3, value=f.field_name)
            ws2.cell(row=r, column=4, value=_fmt(f.extracted_value))
            ws2.cell(row=r, column=5, value=_fmt(f.approved_value))
            ws2.cell(row=r, column=6, value=_fmt(f.confidence))
            ws2.cell(row=r, column=7, value=f.approved_at.isoformat() if f.approved_at else "—")
            for col in range(1, 8):
                ws2.cell(row=r, column=col).border = BORDER
            r += 1

    # Sheet 3 — Audit trail
    ws3 = wb.create_sheet("Audit Trail")
    audit_headers = ["Change ID", "Doc ID", "Field", "Old Value", "New Value", "Changed By", "Changed At"]
    for col, h in enumerate(audit_headers, start=1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
    for col, w in enumerate([10, 8, 20, 28, 28, 12, 22], start=1):
        ws3.column_dimensions[chr(64 + col)].width = w
    for i, a in enumerate(audit_rows, start=2):
        ws3.cell(row=i, column=1, value=a.change_id)
        ws3.cell(row=i, column=2, value=a.doc_id)
        ws3.cell(row=i, column=3, value=a.field_name)
        ws3.cell(row=i, column=4, value=_fmt(a.old_value))
        ws3.cell(row=i, column=5, value=_fmt(a.new_value))
        ws3.cell(row=i, column=6, value=a.changed_by)
        ws3.cell(row=i, column=7, value=a.changed_at.isoformat() if a.changed_at else "—")
        for col in range(1, 8):
            ws3.cell(row=i, column=col).border = BORDER

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(case, documents, fields_by_doc, audit_rows) -> bytes:
    """Build a PDF report for a case and return the bytes."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
        title=f"LogiFlow Case #{case.case_id}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Heading1"],
        textColor=colors.HexColor("#1F4E79"), spaceAfter=12,
    )
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], textColor=colors.HexColor("#1F4E79"))

    story = []
    story.append(Paragraph("LogiFlow — Logistics Paperwork Automation", title_style))
    story.append(Paragraph(f"Case Export — Case #{case.case_id}", h2))
    story.append(Spacer(1, 6))

    summary = [
        ["Customer", _fmt(case.customer)],
        ["Service Type", _fmt(case.service_type)],
        ["Status", _fmt(case.status)],
        ["Created At", case.created_at.isoformat() if case.created_at else "—"],
        ["Documents", str(len(documents))],
        ["Generated At", datetime.utcnow().isoformat(timespec="seconds") + "Z"],
    ]
    t = Table(summary, colWidths=[4 * cm, 12 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EFF3FA")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1F4E79")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(t)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Approved Fields", h2))
    field_rows = [["Doc", "Field", "Approved Value", "Confidence"]]
    for d in documents:
        for f in fields_by_doc.get(d.doc_id, []):
            field_rows.append([
                str(d.doc_id),
                f.field_name,
                _fmt(f.approved_value or f.extracted_value),
                _fmt(f.confidence),
            ])
    if len(field_rows) == 1:
        field_rows.append(["—", "No approved fields", "—", "—"])

    ft = Table(field_rows, colWidths=[1.5 * cm, 4 * cm, 8 * cm, 2.5 * cm])
    ft.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FD")]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
    ]))
    story.append(ft)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Audit Trail", h2))
    audit_data = [["Field", "Old", "New", "User", "When"]]
    for a in audit_rows[:50]:
        audit_data.append([
            a.field_name,
            _fmt(a.old_value),
            _fmt(a.new_value),
            str(a.changed_by) if a.changed_by else "—",
            a.changed_at.strftime("%Y-%m-%d %H:%M") if a.changed_at else "—",
        ])
    if len(audit_data) == 1:
        audit_data.append(["—", "No changes recorded", "—", "—", "—"])

    at = Table(audit_data, colWidths=[3.5 * cm, 4 * cm, 4 * cm, 1.5 * cm, 3 * cm])
    at.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F8FD")]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFBFBF")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BFBFBF")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(at)

    doc.build(story)
    return buf.getvalue()
